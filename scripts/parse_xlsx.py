"""parse_xlsx.py — read the 27-sheet source workbook into clean per-table CSVs.

Every data sheet is padded to a fixed 26-column grid in the source file;
this trims the trailing all-empty columns/rows and writes one CSV per real
table to data/processed/. Documentation sheets (SUMMARY, README, SCHEMA,
FEATURE_POLICY, VIEW_SPEC) are captured as structured text, not loaded as
data tables — they are the ground truth for the data dictionary.
"""
from __future__ import annotations
import json
import sys
from pathlib import Path

import openpyxl
import pandas as pd

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from scripts.config import RAW_XLSX, PROCESSED_DIR, DATA_SHEETS, DOC_SHEETS


def _sheet_to_df(ws) -> pd.DataFrame:
    rows = ws.iter_rows(values_only=True)
    header = list(next(rows))
    # trim trailing None columns from the fixed 26-col grid
    while header and header[-1] is None:
        header.pop()
    ncols = len(header)
    data = []
    for row in rows:
        row = list(row[:ncols])
        if all(v is None for v in row):
            continue
        data.append(row)
    return pd.DataFrame(data, columns=header)


def _sheet_to_records(ws) -> list[list]:
    """For documentation sheets: return non-empty rows as lists of non-None cells."""
    records = []
    for row in ws.iter_rows(values_only=True):
        vals = [v for v in row if v is not None]
        if vals:
            records.append(vals)
    return records


def parse_workbook() -> dict[str, pd.DataFrame]:
    PROCESSED_DIR.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.load_workbook(RAW_XLSX, read_only=True, data_only=True)

    tables: dict[str, pd.DataFrame] = {}
    for name in DATA_SHEETS:
        df = _sheet_to_df(wb[name])
        df.to_csv(PROCESSED_DIR / f"{name.lower()}.csv", index=False)
        tables[name] = df
        print(f"  {name:<22} {len(df):>7,} rows x {len(df.columns)} cols")

    docs = {name: _sheet_to_records(wb[name]) for name in DOC_SHEETS}
    with open(PROCESSED_DIR / "documentation_sheets.json", "w") as fh:
        json.dump(docs, fh, indent=2, default=str)
    print(f"  documentation sheets captured: {', '.join(DOC_SHEETS)}")

    return tables


if __name__ == "__main__":
    print("Parsing workbook ->", RAW_XLSX.name)
    parse_workbook()
